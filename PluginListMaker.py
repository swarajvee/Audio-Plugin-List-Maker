import os
import openpyxl
from datetime import datetime
import numpy as np

def ExcelMaker(vst_list, au_list, vst3_list, aax_list):
    """Makes excel file"""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'My Plugin List'
    headers = ['Sl.No', 'VST3 Plugins', 'VST Plugins', 'AU Plugins', 'AAX Plugins']
    sheet.append(headers)

    sl_no = np.arange(1, len(max(vst3_list, vst_list, au_list, aax_list, key=len)) + 1, 1)
    for o, no in enumerate(sl_no, start=2):
        sheet.cell(row=o, column=1, value=no)

    new_vst3_list = []
    new_vst_list = []
    new_au_list = []
    new_aax_list = []

    if vst3_list:
        new_vst3_list = [vst3.replace('.vst3', '') for vst3 in vst3_list]
        for i, vst3 in enumerate(new_vst3_list, start=2):
            sheet.cell(row=i, column=2, value=vst3)

    if vst_list:
        new_vst_list = [vst.replace('.vst', '') for vst in vst_list]
        for j, vst in enumerate(new_vst_list, start=2):
            sheet.cell(row=j, column=3, value=vst)

    if au_list:
        new_au_list = [au.replace('.component', '') for au in au_list]
        for k, au in enumerate(new_au_list, start=2):
            sheet.cell(row=k, column=4, value=au)

    if aax_list:
        new_aax_list = [aax.replace('.aaxplugin', '') for aax in aax_list]
        for l, aax in enumerate(new_aax_list, start=2):
            sheet.cell(row=l, column=5, value=aax)

    sheet.row_dimensions[1].height = 48
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 35
    sheet.column_dimensions['D'].width = 35
    sheet.column_dimensions['E'].width = 40

    common_plugins_sheet = wb.create_sheet(title='Common Plugins')
    headers_common = ['Sl. No', 'Common Plugins', 'Versions']
    common_plugins_sheet.append(headers_common)

    combined_plugins = sorted(set(new_vst3_list) | set(new_vst_list) | set(new_au_list) | set(new_aax_list))
    common_plugins = []
    plugin_versions = []

    for plugin in combined_plugins:
        versions = []
        if plugin in new_vst3_list:
            versions.append('VST3')
        if plugin in new_vst_list:
            versions.append('VST')
        if plugin in new_au_list:
            versions.append('AU')
        if plugin in new_aax_list:
            versions.append('AAX')
        if len(versions) > 1:
            common_plugins.append(plugin)
            plugin_versions.append(','.join(versions))

    sl_no_common = np.arange(1, len(common_plugins) + 1, 1)
    for p, no in enumerate(sl_no_common, start=2):
        common_plugins_sheet.cell(row=p, column=1, value=no)

    for m, (common_plugin, version) in enumerate(zip(common_plugins, plugin_versions), start=2):
        common_plugins_sheet.cell(row=m, column=2, value=common_plugin)
        common_plugins_sheet.cell(row=m, column=3, value=version)

    common_plugins_sheet.row_dimensions[1].height = 48
    common_plugins_sheet.column_dimensions['B'].width = 40
    common_plugins_sheet.column_dimensions['C'].width = 20

    wb.save(filename=f"plugin list_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")

def find_plugins(plugin_path, file_extension):
    plugin_list = []
    
    def search_plugins(path):
        for plugin in os.listdir(path):
            plugin_path = os.path.join(path, plugin)
            if plugin.endswith(file_extension):
                plugin_list.append(plugin)
            elif os.path.isdir(plugin_path):
                search_plugins(plugin_path)
    
    if os.path.isdir(plugin_path):
        search_plugins(plugin_path)
    
    return plugin_list

def PluginListMaker(vst_path, au_path, vst3_path, aax_path):

print("Please enter the plugin directory paths (press enter if no path)")
vst_path = input("Enter the VST plugin directory path: ")
vst3_path = input("Enter the VST3 plugin directory path: ")
au_path = input("Enter the AU plugin directory path: ")
aax_path = input("Enter the AAX plugin directory path: ")

PluginListMaker(vst_path, au_path, vst3_path, aax_path)

print(f"Excel file is created in {os.getcwd()}")

